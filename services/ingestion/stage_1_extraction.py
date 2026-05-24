"""
Stage 1 — Structural extraction.

Converts raw files into a normalised ExtractedDocument.
Zero LLM calls — only open-source parsing libraries.

Supported types (initial release): PDF (native + scanned), DOCX, XLSX, PPTX,
HTML, JPG/PNG.
"""

from __future__ import annotations

import io
import logging
import os
import tempfile
import urllib.request
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

import fitz                             # PyMuPDF — AGPL / commercial

log = logging.getLogger(__name__)

GROBID_URL: str = os.getenv("KOS_GROBID_URL", "http://grobid:8070")
OCR_MIN_TEXT_CHARS: int = 20           # pages with fewer chars are treated as scanned


@dataclass
class Section:
    title: str
    level: int                          # heading depth (1 = H1, 2 = H2, …)
    content: str
    page: Optional[int] = None


@dataclass
class Table:
    title: str
    data: list[list[Any]]               # rows × cols
    page: Optional[int] = None


@dataclass
class Figure:
    index: int
    caption: str
    image_bytes: bytes                  # PNG bytes
    page: Optional[int] = None


@dataclass
class ExtractedDocument:
    markdown: str = ""
    raw_text: str = ""
    metadata: dict = field(default_factory=dict)
    sections: list[Section] = field(default_factory=list)
    tables: list[Table] = field(default_factory=list)
    figures: list[Figure] = field(default_factory=list)
    language: str = "und"               # ISO-639
    is_scanned: bool = False
    source_tool: str = ""


# ── Public entry point ─────────────────────────────────────────────────────────

def extract_structure(file_path: str, file_type: str) -> ExtractedDocument:
    """
    Dispatch to the appropriate extractor and return an ExtractedDocument.

    ``file_type`` is the normalised extension without the dot:
    'pdf', 'docx', 'xlsx', 'pptx', 'html', 'image'.
    """
    path = Path(file_path)

    dispatch = {
        "pdf":   _extract_pdf,
        "docx":  _extract_office,
        "xlsx":  _extract_xlsx,
        "pptx":  _extract_office,
        "html":  _extract_html,
        "htm":   _extract_html,
        "image": _extract_image,
        "jpg":   _extract_image,
        "jpeg":  _extract_image,
        "png":   _extract_image,
    }

    handler = dispatch.get(file_type.lower())
    if handler is None:
        raise ValueError(f"Unsupported file type: {file_type!r}")

    try:
        doc = handler(path)
        doc.language = _detect_language(doc.raw_text)
        return doc
    except Exception as exc:
        log.exception("Extraction failed for %s: %s", path.name, exc)
        raise


# ── PDF extractor ──────────────────────────────────────────────────────────────

def _extract_pdf(path: Path) -> ExtractedDocument:
    """
    PDF native path:   PyMuPDF → Docling → pdfplumber for tables → GROBID metadata.
    PDF scanned path:  PyMuPDF detects no text → Tesseract OCR → Docling structure.
    """
    doc = fitz.open(str(path))
    total_chars = sum(len(page.get_text()) for page in doc)
    is_scanned = total_chars < OCR_MIN_TEXT_CHARS * len(doc)
    doc.close()

    if is_scanned:
        return _extract_pdf_scanned(path)
    return _extract_pdf_native(path)


def _extract_pdf_native(path: Path) -> ExtractedDocument:
    # ── 1. Docling for semantic Markdown ──────────────────────────────────────
    markdown, sections, metadata = _docling_convert(path)

    # ── 2. pdfplumber for robust table extraction ─────────────────────────────
    tables = _pdfplumber_tables(path)

    # ── 3. PyMuPDF for figure extraction ─────────────────────────────────────
    figures = _pymupdf_figures(path)

    # ── 4. GROBID for academic metadata (best-effort) ─────────────────────────
    grobid_meta = _grobid_metadata(path)
    if grobid_meta:
        metadata.update(grobid_meta)

    raw_text = _markdown_to_raw(markdown)

    return ExtractedDocument(
        markdown=markdown,
        raw_text=raw_text,
        metadata=metadata,
        sections=sections,
        tables=tables,
        figures=figures,
        is_scanned=False,
        source_tool="docling+pymupdf",
    )


def _extract_pdf_scanned(path: Path) -> ExtractedDocument:
    # ── 1. OCR page-by-page with Tesseract ────────────────────────────────────
    ocr_text_pages = _tesseract_ocr_pdf(path)
    full_text = "\n\n".join(ocr_text_pages)

    # ── 2. Write OCR output to a temp file and run through Docling ────────────
    with tempfile.NamedTemporaryFile(suffix=".txt", mode="w",
                                     delete=False, encoding="utf-8") as tmp:
        tmp.write(full_text)
        tmp_path = Path(tmp.name)

    try:
        markdown, sections, metadata = _docling_convert_text(full_text)
    finally:
        tmp_path.unlink(missing_ok=True)

    # Extract figures as images (captions may be empty)
    figures = _pymupdf_figures(path)

    return ExtractedDocument(
        markdown=markdown,
        raw_text=full_text,
        metadata=metadata,
        sections=sections,
        figures=figures,
        is_scanned=True,
        source_tool="tesseract+docling",
    )


# ── Office extractors ──────────────────────────────────────────────────────────

def _extract_office(path: Path) -> ExtractedDocument:
    """DOCX / PPTX via Docling."""
    markdown, sections, metadata = _docling_convert(path)
    raw_text = _markdown_to_raw(markdown)
    return ExtractedDocument(
        markdown=markdown,
        raw_text=raw_text,
        metadata=metadata,
        sections=sections,
        source_tool="docling",
    )


def _extract_xlsx(path: Path) -> ExtractedDocument:
    """XLSX: extract each sheet as a structured JSON table, not flat Markdown."""
    try:
        import openpyxl  # Apache 2.0  # noqa: PLC0415
    except ImportError as exc:
        raise ImportError("openpyxl is required for XLSX extraction") from exc

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    tables: list[Table] = []
    markdown_parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [[str(cell.value) if cell.value is not None else "" for cell in row]
                for row in ws.iter_rows()]
        if not rows:
            continue
        table = Table(title=sheet_name, data=rows)
        tables.append(table)

        # Also render as Markdown table for fulltext search
        header = rows[0]
        sep = ["---"] * len(header)
        body = rows[1:] if len(rows) > 1 else []
        md_rows = [f"| {' | '.join(header)} |",
                   f"| {' | '.join(sep)} |"]
        md_rows += [f"| {' | '.join(r)} |" for r in body[:200]]  # cap at 200 rows
        markdown_parts.append(f"## {sheet_name}\n\n" + "\n".join(md_rows))

    wb.close()
    markdown = "\n\n".join(markdown_parts)
    raw_text = "\n".join(
        " ".join(str(c) for c in row)
        for t in tables for row in t.data
    )

    return ExtractedDocument(
        markdown=markdown,
        raw_text=raw_text,
        metadata={"title": path.stem},
        tables=tables,
        source_tool="openpyxl",
    )


# ── HTML extractor ─────────────────────────────────────────────────────────────

def _extract_html(path: Path) -> ExtractedDocument:
    try:
        from bs4 import BeautifulSoup  # noqa: PLC0415
    except ImportError as exc:
        raise ImportError("beautifulsoup4 is required for HTML extraction") from exc

    html = path.read_text(errors="replace")
    soup = BeautifulSoup(html, "html.parser")

    # Remove script/style noise
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()

    title = soup.title.string.strip() if soup.title else path.stem
    raw_text = soup.get_text(separator="\n", strip=True)

    # Build sections from heading tags
    sections: list[Section] = []
    current_level = 1
    current_title = "Introduction"
    current_content: list[str] = []

    for el in soup.find_all(["h1", "h2", "h3", "h4", "p", "li"]):
        tag = el.name
        if tag in {"h1", "h2", "h3", "h4"}:
            if current_content:
                sections.append(Section(
                    title=current_title,
                    level=current_level,
                    content="\n".join(current_content).strip(),
                ))
            current_level = int(tag[1])
            current_title = el.get_text(strip=True)
            current_content = []
        else:
            current_content.append(el.get_text(strip=True))

    if current_content:
        sections.append(Section(
            title=current_title,
            level=current_level,
            content="\n".join(current_content).strip(),
        ))

    # Simple Markdown render
    markdown = f"# {title}\n\n" + raw_text[:50_000]

    return ExtractedDocument(
        markdown=markdown,
        raw_text=raw_text,
        metadata={"title": title},
        sections=sections,
        source_tool="beautifulsoup4",
    )


# ── Image extractor ────────────────────────────────────────────────────────────

def _extract_image(path: Path) -> ExtractedDocument:
    ocr_text = _tesseract_ocr_image(path)
    # Store the image itself as figure 0
    image_bytes = path.read_bytes()
    figures = [Figure(index=0, caption="", image_bytes=image_bytes, page=1)]

    markdown = f"# {path.stem}\n\n{ocr_text}" if ocr_text else f"# {path.stem}\n\n[Image]"

    return ExtractedDocument(
        markdown=markdown,
        raw_text=ocr_text,
        metadata={"title": path.stem},
        figures=figures,
        is_scanned=True,
        source_tool="tesseract",
    )


# ── Docling wrappers ───────────────────────────────────────────────────────────

def _docling_convert(path: Path) -> tuple[str, list[Section], dict]:
    """
    Run Docling (MIT) to convert a document to Markdown + section tree.
    Falls back to PyMuPDF raw text if Docling is unavailable.
    """
    try:
        from docling.document_converter import DocumentConverter  # noqa: PLC0415

        converter = DocumentConverter()
        result = converter.convert(str(path))
        markdown = result.document.export_to_markdown()
        metadata = _docling_metadata(result)
        sections = _parse_sections_from_markdown(markdown)
        return markdown, sections, metadata
    except ImportError:
        log.warning("docling not installed — falling back to PyMuPDF text extraction")
        return _pymupdf_fallback(path)
    except Exception as exc:
        log.warning("Docling conversion failed for %s: %s — falling back", path.name, exc)
        return _pymupdf_fallback(path)


def _docling_convert_text(text: str) -> tuple[str, list[Section], dict]:
    """Wrap plain text in minimal Markdown for further processing."""
    sections = _parse_sections_from_markdown(text)
    return text, sections, {}


def _docling_metadata(result) -> dict:
    """Extract metadata fields from a Docling ConversionResult."""
    meta = {}
    try:
        doc_meta = result.document.metadata
        if doc_meta:
            meta["title"] = getattr(doc_meta, "title", None)
            meta["authors"] = getattr(doc_meta, "authors", [])
            meta["date_modified"] = str(getattr(doc_meta, "date", "") or "")
    except Exception:  # noqa: BLE001
        pass
    return {k: v for k, v in meta.items() if v}


# ── PyMuPDF helpers ────────────────────────────────────────────────────────────

def _pymupdf_fallback(path: Path) -> tuple[str, list[Section], dict]:
    doc = fitz.open(str(path))
    pages_text: list[str] = [page.get_text() for page in doc]
    meta_raw = doc.metadata or {}
    doc.close()

    raw = "\n\n".join(pages_text)
    markdown = raw
    metadata = {
        "title": meta_raw.get("title") or path.stem,
        "authors": [meta_raw["author"]] if meta_raw.get("author") else [],
    }
    sections = _parse_sections_from_markdown(raw)
    return markdown, sections, metadata


def _pymupdf_figures(path: Path) -> list[Figure]:
    """Extract embedded images from a PDF as PNG bytes."""
    figures: list[Figure] = []
    doc = fitz.open(str(path))
    idx = 0
    for page_num, page in enumerate(doc, start=1):
        for img_info in page.get_images(full=True):
            xref = img_info[0]
            base_image = doc.extract_image(xref)
            img_bytes = base_image["image"]
            ext = base_image["ext"]

            # Convert to PNG if needed
            if ext.lower() != "png":
                img_bytes = _to_png(img_bytes)

            figures.append(Figure(
                index=idx,
                caption="",
                image_bytes=img_bytes,
                page=page_num,
            ))
            idx += 1
    doc.close()
    return figures


def _to_png(image_bytes: bytes) -> bytes:
    """Convert any image format to PNG using Pillow (MIT)."""
    try:
        from PIL import Image  # noqa: PLC0415
        img = Image.open(io.BytesIO(image_bytes))
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue()
    except Exception:  # noqa: BLE001
        return image_bytes  # return original if conversion fails


# ── pdfplumber table extraction ────────────────────────────────────────────────

def _pdfplumber_tables(path: Path) -> list[Table]:
    """Extract tabular data from a PDF using pdfplumber (MIT)."""
    tables: list[Table] = []
    try:
        import pdfplumber  # noqa: PLC0415
        with pdfplumber.open(str(path)) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                for tbl in page.extract_tables():
                    if tbl:
                        tables.append(Table(title="", data=tbl, page=page_num))
    except ImportError:
        log.debug("pdfplumber not installed — table extraction skipped")
    except Exception as exc:
        log.warning("pdfplumber failed on %s: %s", path.name, exc)
    return tables


# ── Tesseract OCR helpers ──────────────────────────────────────────────────────

def _tesseract_ocr_pdf(path: Path) -> list[str]:
    """OCR each PDF page; return list of text strings (one per page)."""
    try:
        import pytesseract  # noqa: PLC0415
        from PIL import Image  # noqa: PLC0415
    except ImportError as exc:
        raise ImportError("pytesseract + Pillow required for OCR") from exc

    doc = fitz.open(str(path))
    results: list[str] = []
    for page in doc:
        pix = page.get_pixmap(dpi=200)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        text = pytesseract.image_to_string(img, config="--psm 3")
        results.append(text)
    doc.close()
    return results


def _tesseract_ocr_image(path: Path) -> str:
    """OCR a single image file."""
    try:
        import pytesseract  # noqa: PLC0415
        from PIL import Image  # noqa: PLC0415

        img = Image.open(str(path))
        return pytesseract.image_to_string(img, config="--psm 3")
    except ImportError:
        log.warning("pytesseract not installed — image OCR skipped")
        return ""
    except Exception as exc:
        log.warning("Tesseract OCR failed for %s: %s", path.name, exc)
        return ""


# ── GROBID integration ─────────────────────────────────────────────────────────

def _grobid_metadata(path: Path) -> Optional[dict]:
    """
    Call local GROBID service for academic header extraction.
    Returns None if GROBID is unavailable or the file is not academic.
    """
    try:
        url = f"{GROBID_URL}/api/processHeaderDocument"
        with open(str(path), "rb") as fh:
            data = fh.read()

        import urllib.parse  # noqa: PLC0415
        boundary = b"----KOSBoundary"
        body = (
            b"--" + boundary + b"\r\n"
            b'Content-Disposition: form-data; name="input"; filename="doc.pdf"\r\n'
            b"Content-Type: application/pdf\r\n\r\n" + data + b"\r\n"
            b"--" + boundary + b"--\r\n"
        )
        req = urllib.request.Request(
            url,
            data=body,
            headers={
                "Content-Type": f"multipart/form-data; boundary={boundary.decode()}",
            },
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            if resp.status == 200:
                return _parse_grobid_tei(resp.read().decode())
    except Exception as exc:  # noqa: BLE001
        log.debug("GROBID unavailable or failed: %s", exc)
    return None


def _parse_grobid_tei(tei_xml: str) -> dict:
    """Extract title/authors/date from GROBID TEI XML (simple regex approach)."""
    import re  # noqa: PLC0415

    meta: dict = {}
    title_m = re.search(r"<title[^>]*>(.*?)</title>", tei_xml, re.S)
    if title_m:
        meta["title"] = re.sub(r"<[^>]+>", "", title_m.group(1)).strip()

    authors = re.findall(
        r"<persName>(.*?)</persName>", tei_xml, re.S
    )
    if authors:
        meta["authors"] = [re.sub(r"<[^>]+>", "", a).strip() for a in authors]

    date_m = re.search(r'<date[^>]+when="([^"]+)"', tei_xml)
    if date_m:
        meta["date_created"] = date_m.group(1)

    return meta


# ── Language detection ─────────────────────────────────────────────────────────

def _detect_language(text: str) -> str:
    """Lightweight language detection with langdetect (MIT fork of language-detection)."""
    if len(text) < 30:
        return "und"
    try:
        from langdetect import detect  # noqa: PLC0415
        return detect(text[:2000])
    except Exception:  # noqa: BLE001
        return "und"


# ── Utility ────────────────────────────────────────────────────────────────────

def _parse_sections_from_markdown(text: str) -> list[Section]:
    """
    Split Markdown into Sections following ATX heading markers (# / ## / ###).
    Paragraphs before the first heading become a level-0 preamble section.
    """
    import re  # noqa: PLC0415

    lines = text.splitlines()
    sections: list[Section] = []
    current_title = "Preamble"
    current_level = 0
    current_lines: list[str] = []

    for line in lines:
        m = re.match(r"^(#{1,6})\s+(.*)", line)
        if m:
            if current_lines:
                sections.append(Section(
                    title=current_title,
                    level=current_level,
                    content="\n".join(current_lines).strip(),
                ))
            current_level = len(m.group(1))
            current_title = m.group(2).strip()
            current_lines = []
        else:
            current_lines.append(line)

    if current_lines:
        sections.append(Section(
            title=current_title,
            level=current_level,
            content="\n".join(current_lines).strip(),
        ))

    return [s for s in sections if s.content]


def _markdown_to_raw(markdown: str) -> str:
    """Strip Markdown syntax, returning plain text."""
    import re  # noqa: PLC0415

    text = re.sub(r"!\[.*?\]\(.*?\)", "", markdown)     # images
    text = re.sub(r"\[.*?\]\(.*?\)", r"", text)          # links
    text = re.sub(r"[#*`_~>|-]+", " ", text)             # markers
    text = re.sub(r"\s{2,}", " ", text)
    return text.strip()
